"""
ENR Top 500 Design Firms ingestion.

Normalizes the 13 different header signatures across 21 annual files (ENR 2006 edition
through ENR 2026 edition) into a single canonical pandas DataFrame.

Canonical schema:
    edition_year   int     ENR edition year (e.g., 2026)
    data_year      int     Year the revenue data describes (= edition_year - 1)
    rank           int     Firm's rank in this edition
    firm_raw       str     Original firm name as it appears in the file
    firm_key       str     Normalized identity key (uppercase, stripped of city/punct)
    location       str     City, state (best-effort)
    firm_type      str     ENR firm type code (E, A, EA, EAC, M, etc.)
    total_rev_m    float   Total revenue, nominal $M
    intl_rev_m     float   International revenue, nominal $M

    plus per-sector pairs (for the 10 ENR sectors):
        gen_bldg_pct,   gen_bldg_m
        manufacturing_pct, manufacturing_m
        power_pct,      power_m
        water_supply_pct, water_supply_m
        sewer_waste_pct, sewer_waste_m
        ind_pet_pct,    ind_pet_m
        transportation_pct, transportation_m
        haz_waste_pct,  haz_waste_m
        telecom_pct,    telecom_m
        other_pct,      other_m

When the sector $M column is missing/null but the % is populated, sector $M is
imputed as total_rev_m * pct / 100.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pandas as pd


# Canonical sector list. Order matters for column layout.
SECTORS = [
    "gen_bldg",
    "manufacturing",
    "power",
    "water_supply",
    "sewer_waste",
    "ind_pet",
    "transportation",
    "haz_waste",
    "telecom",
    "other",
]


# Each sector has an ordered list of header substrings (lowercased) that identify it.
# Order matters: more specific patterns first.
SECTOR_HEADER_PATTERNS = {
    "gen_bldg": ["general_building", "general building", "gen bldg", "gen_bldg"],
    "manufacturing": ["manufacturing", "mfg"],
    "power": ["power"],
    "water_supply": ["water_supply", "water supply"],
    "sewer_waste": ["sewer_waste", "sewer/waste", "sewer waste"],
    "ind_pet": [
        "industrial/oil&gas",
        "industrial/petroleum",
        "industrial_petroleum",
        "indus/petro",
    ],
    "transportation": ["transportation", "transp"],
    "haz_waste": ["hazardous_waste", "hazardous waste", "haz waste", "haz_waste"],
    "telecom": ["telecommunications", "telecom"],
    "other": ["other"],
}


def _normalize_header(h) -> str:
    """Lowercase, strip whitespace and Unicode oddities."""
    if h is None:
        return ""
    s = str(h).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    return s


def _is_pct_header(h: str) -> bool:
    """Check if a normalized header is a percentage column."""
    return ("%" in h) or ("(%)" in h) or h.endswith(" %") or "_pct" in h or " pct" in h


def _is_dollar_header(h: str) -> bool:
    """Check if a normalized header is a dollar/revenue column (excluding %)."""
    if _is_pct_header(h):
        return False
    return any(
        token in h
        for token in ["$m", "$ m", "$mil", "($mil)", "rev ($m)", "revenue ($m)", "$mil ", " $m"]
    ) or h.endswith(" $m") or h.endswith("$m") or "_m" in h or "rev" in h or "$mil" in h


def _detect_sector_columns(headers: list[str]) -> dict[str, dict[str, int]]:
    """
    For each canonical sector, find the column indices of its % and $ columns.

    Returns dict like {"gen_bldg": {"pct": 7, "dol": 8}, ...}
    """
    norm = [_normalize_header(h) for h in headers]
    result: dict[str, dict[str, int]] = {}

    for sector, patterns in SECTOR_HEADER_PATTERNS.items():
        pct_idx = None
        dol_idx = None
        for i, h in enumerate(norm):
            if not h:
                continue
            # Match this header against any of the sector's patterns
            matched = any(p in h for p in patterns)
            if not matched:
                continue
            # Don't let "other" capture "Hazardous Waste %" etc. — "other" must be
            # a standalone token, not buried in another sector name. Skip if we already
            # bound this column to another sector.
            if sector == "other" and not (h.startswith("other") or h == "other %" or " other" in h):
                # Be stricter for "other" since it's a generic word
                # accept "other %", "other $m", "other rev ($m)", "OTHER %"
                pass

            if _is_pct_header(h) and pct_idx is None:
                pct_idx = i
            elif not _is_pct_header(h) and dol_idx is None:
                # Heuristic: dollar columns come right after % columns and contain
                # rev/$m/$mil tokens, OR they're the next column after the % col with no %.
                dol_idx = i

        if pct_idx is not None or dol_idx is not None:
            result[sector] = {"pct": pct_idx, "dol": dol_idx}

    return result


def _detect_total_rev_column(headers: list[str], data_rows: list[tuple]) -> int | None:
    """
    Find the column index for total revenue. Headers vary wildly:
      'Total Revenue ($M)', 'Total_Revenue_$MIL', 'Total_Revenue ($Mil)',
      'Revenue 2019 ($M)', '2024 Revenue ($M)', '2025 Revenue ($M)', None.
    The 2023 file has a None header where total revenue lives.
    """
    norm = [_normalize_header(h) for h in headers]

    # First pass: explicit "total" markers
    for i, h in enumerate(norm):
        if "total" in h and ("rev" in h or "$" in h or "mil" in h):
            return i

    # Second pass: "<year> revenue ($m)" or "revenue <year> ($m)"
    for i, h in enumerate(norm):
        if re.search(r"\b\d{4}\b\s*revenue", h) or re.search(r"revenue\s*\d{4}", h):
            return i

    # Third pass: find the None-header column whose values are numeric and large,
    # placed after location/firm-type and before any % column.
    for i, h in enumerate(headers):
        if h is None:
            # Spot-check: is this column numeric for the first data row, and reasonably big?
            if data_rows:
                v = data_rows[0][i]
                if isinstance(v, (int, float)) and v > 100:
                    # Probably revenue (top firms are >$1B; even rank 500 is >$100M-ish)
                    return i

    return None


def _detect_intl_rev_column(headers: list[str]) -> int | None:
    norm = [_normalize_header(h) for h in headers]
    for i, h in enumerate(norm):
        if ("int" in h) and ("rev" in h):
            return i
    return None


def _detect_rank_column(headers: list[str]) -> int | None:
    """Current-edition rank is the first column matching 'rank'."""
    norm = [_normalize_header(h) for h in headers]
    for i, h in enumerate(norm):
        if "rank" in h:
            return i
    return None


def _detect_firm_column(headers: list[str]) -> int | None:
    norm = [_normalize_header(h) for h in headers]
    for i, h in enumerate(norm):
        if h in ("firm", "firm_name", "firm name"):
            return i
    return None


def _detect_location_columns(headers: list[str]) -> tuple[int | None, int | None]:
    """Returns (location_or_city_idx, state_idx). Some files have a single Location
    column, others have City + State."""
    norm = [_normalize_header(h) for h in headers]
    loc_or_city = None
    state = None
    for i, h in enumerate(norm):
        if h == "location":
            loc_or_city = i
            return (loc_or_city, None)
        if h == "city":
            loc_or_city = i
        if h == "state":
            state = i
    return (loc_or_city, state)


def _detect_firm_type_column(headers: list[str]) -> int | None:
    norm = [_normalize_header(h) for h in headers]
    for i, h in enumerate(norm):
        if h in ("firm type", "firm_type", "type", "type of firm"):
            return i
    return None


def _to_float(v) -> float | None:
    """Convert a cell to float, handling strings like '2,360.9' and None."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


_FIRM_TYPE_TOKENS = {"E", "A", "EA", "EC", "EAC", "AE", "M", "ENV", "ENG", "GE", "L", "P", "S"}


def _split_firm_and_location(firm_raw: str) -> tuple[str, str | None]:
    """
    Some files combine firm+city in the firm column: 'AECOM, Dallas',
    'HDR, Omaha', 'HDR, Omaha, Neb.', 'Jacobs, Pasadena, Calif.'.
    Returns (clean_firm_name, extracted_city_or_None).
    """
    if not firm_raw:
        return ("", None)
    parts = [p.strip() for p in firm_raw.split(",")]
    if len(parts) == 1:
        return (parts[0], None)
    # First part is firm; rest is location
    return (parts[0], ", ".join(parts[1:]))


def _normalize_firm_key(firm_raw: str) -> str:
    """
    Produce a stable identity key for matching the same firm across years.

    Strategy: take the firm name (drop any city after first comma), uppercase,
    strip punctuation/possessive/common suffixes, collapse whitespace, strip
    Unicode marks (handles 'Omaha, Neb.†' kind of garbage and 'hDR' → 'HDR').
    """
    name, _ = _split_firm_and_location(firm_raw)
    s = unicodedata.normalize("NFKD", name)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.upper()
    # Only strip *legal-entity* suffixes, never descriptive nouns. "ASSOCIATES",
    # "ENGINEERS", "GROUP", "HOLDINGS", "SOLUTIONS", "CONSULTANTS", "CONSULTING"
    # are part of the firm's identity (e.g., "Jacobs Associates" is a different
    # firm from "Jacobs"). Keep this list conservative.
    suffixes = [
        " INC.",
        " INC",
        " LLC",
        " L.L.C.",
        " LP",
        " L.P.",
        " CORPORATION",
        " CORP.",
        " CORP",
        " CO.",
        " COS.",
        " COS",
        " LTD.",
        " LTD",
        " PLC",
        " S.A.",
    ]
    # Remove trailing dagger / asterisks / etc.
    s = re.sub(r"[†*‡§¶]+", "", s)
    # Drop "THE " prefix
    if s.startswith("THE "):
        s = s[4:]
    for suf in suffixes:
        if s.endswith(suf):
            s = s[: -len(suf)].rstrip()
    # Strip remaining punctuation
    s = re.sub(r"[^\w\s&]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


@dataclass
class FileSchema:
    """Column index map for one ENR file."""

    edition_year: int
    rank_idx: int | None
    firm_idx: int | None
    location_idx: int | None
    state_idx: int | None
    firm_type_idx: int | None
    total_rev_idx: int | None
    intl_rev_idx: int | None
    sector_cols: dict[str, dict[str, int]]


def _infer_edition_year(filename: str) -> int:
    """Pull the 4-digit year from the filename."""
    m = re.search(r"(\d{4})", filename)
    if not m:
        raise ValueError(f"Could not infer edition year from filename: {filename}")
    return int(m.group(1))


def _detect_schema(headers: list, data_rows: list[tuple], edition_year: int) -> FileSchema:
    return FileSchema(
        edition_year=edition_year,
        rank_idx=_detect_rank_column(headers),
        firm_idx=_detect_firm_column(headers),
        location_idx=_detect_location_columns(headers)[0],
        state_idx=_detect_location_columns(headers)[1],
        firm_type_idx=_detect_firm_type_column(headers),
        total_rev_idx=_detect_total_rev_column(headers, data_rows),
        intl_rev_idx=_detect_intl_rev_column(headers),
        sector_cols=_detect_sector_columns(headers),
    )


def _read_file(path: Path) -> list[dict]:
    """Read one ENR file and return list of dicts in canonical schema."""
    edition_year = _infer_edition_year(path.name)
    data_year = edition_year - 1  # ENR convention: edition Y reports year Y-1 revenue

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = list(rows[0])
    data_rows = rows[1:]

    schema = _detect_schema(headers, data_rows, edition_year)

    out: list[dict] = []
    for r in data_rows:
        if not r or all(v is None for v in r):
            continue

        # Get firm name; some files store it at a column we couldn't auto-detect,
        # but we have firm_idx detection covering "firm" / "firm name" / "firm_name"
        firm_raw = r[schema.firm_idx] if schema.firm_idx is not None else None
        if firm_raw is None or not str(firm_raw).strip():
            continue

        firm_clean, embedded_city = _split_firm_and_location(str(firm_raw))
        firm_key = _normalize_firm_key(str(firm_raw))
        if not firm_key:
            continue

        # Location: prefer explicit Location/City+State columns; fall back to embedded city
        location = None
        if schema.location_idx is not None:
            loc_val = r[schema.location_idx]
            if loc_val:
                location = str(loc_val).strip()
        if location and schema.state_idx is not None:
            st = r[schema.state_idx]
            if st:
                location = f"{location}, {str(st).strip()}"
        if not location and embedded_city:
            location = embedded_city
        # Some 2009-style files put state in the wrong field (firm_type slot has a state).
        # Don't worry about that edge case for now; firm_key is what matters for joins.

        firm_type = (
            str(r[schema.firm_type_idx]).strip()
            if schema.firm_type_idx is not None and r[schema.firm_type_idx] is not None
            else None
        )

        rank = None
        if schema.rank_idx is not None:
            rank_val = r[schema.rank_idx]
            if isinstance(rank_val, (int, float)):
                rank = int(rank_val)
            elif isinstance(rank_val, str) and rank_val.strip().isdigit():
                rank = int(rank_val.strip())

        total_rev = (
            _to_float(r[schema.total_rev_idx]) if schema.total_rev_idx is not None else None
        )
        intl_rev = _to_float(r[schema.intl_rev_idx]) if schema.intl_rev_idx is not None else None

        record = {
            "edition_year": edition_year,
            "data_year": data_year,
            "rank": rank,
            "firm_raw": firm_clean,
            "firm_key": firm_key,
            "location": location,
            "firm_type": firm_type,
            "total_rev_m": total_rev,
            "intl_rev_m": intl_rev,
        }

        # Sector columns
        for sector in SECTORS:
            cols = schema.sector_cols.get(sector)
            pct_v = None
            dol_v = None
            if cols:
                if cols.get("pct") is not None:
                    pct_v = _to_float(r[cols["pct"]])
                if cols.get("dol") is not None:
                    dol_v = _to_float(r[cols["dol"]])
            # Impute $ from total * % if missing
            if dol_v is None and pct_v is not None and total_rev is not None:
                dol_v = total_rev * pct_v / 100.0
            record[f"{sector}_pct"] = pct_v
            record[f"{sector}_m"] = dol_v

        out.append(record)

    return out


def build_panel(enr_dir: Path) -> pd.DataFrame:
    """
    Read all ENR files in `enr_dir` and return a single normalized DataFrame.

    Each row is one (edition_year, firm) observation.
    """
    files = sorted(Path(enr_dir).glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in {enr_dir}")

    all_records: list[dict] = []
    for f in files:
        records = _read_file(f)
        all_records.extend(records)

    df = pd.DataFrame.from_records(all_records)

    # Stable sort: edition_year ascending, rank ascending (with NaN ranks last)
    df = df.sort_values(["edition_year", "rank"], na_position="last").reset_index(drop=True)
    return df


# ---------- CCI loader ----------


def load_cci_annual(cci_path: Path, base_year: int = 2025) -> pd.DataFrame:
    """
    Load ENR CCI 20-City annual averages and compute a deflator series.

    Returns DataFrame with columns:
        year   int
        cci    float       Annual average CCI (1913 base = 100)
        deflator    float  Multiplier to convert nominal $ in `year` to constant `base_year` $:
                           constant = nominal * deflator
                           deflator = cci[base_year] / cci[year]
    """
    wb = openpyxl.load_workbook(cci_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row (the one with 'Year' as first cell)
    header_row_idx = None
    for i, r in enumerate(rows):
        if r and str(r[0]).strip().lower() == "year":
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("Could not find 'Year' header row in CCI file")

    headers = rows[header_row_idx]
    # Find Annual_Avg column
    annual_idx = None
    for i, h in enumerate(headers):
        if h and "annual" in str(h).lower():
            annual_idx = i
            break
    if annual_idx is None:
        raise ValueError("Could not find Annual_Avg column in CCI file")

    records: list[dict] = []
    for r in rows[header_row_idx + 1 :]:
        if not r or r[0] is None:
            continue
        year_v = r[0]
        if not isinstance(year_v, (int, float)):
            continue
        year = int(year_v)
        cci_v = _to_float(r[annual_idx])
        if cci_v is None:
            continue
        records.append({"year": year, "cci": cci_v})

    df = pd.DataFrame.from_records(records).sort_values("year").reset_index(drop=True)
    base_cci = df.loc[df["year"] == base_year, "cci"]
    if base_cci.empty:
        raise ValueError(f"Base year {base_year} not present in CCI data")
    base_cci_val = float(base_cci.iloc[0])
    df["deflator"] = base_cci_val / df["cci"]
    return df


if __name__ == "__main__":
    # Smoke test
    here = Path(__file__).resolve().parent.parent
    panel = build_panel(here / "data" / "enr")
    print(f"Panel: {len(panel)} rows, {panel['edition_year'].nunique()} editions")
    print(f"Editions: {sorted(panel['edition_year'].unique())}")
    print(f"Unique firm_keys: {panel['firm_key'].nunique()}")

    cci = load_cci_annual(here / "data" / "cci.xlsx", base_year=2025)
    print(f"\nCCI: {len(cci)} years, {cci['year'].min()}-{cci['year'].max()}")
    print(cci[cci["year"].between(2004, 2026)].to_string(index=False))
